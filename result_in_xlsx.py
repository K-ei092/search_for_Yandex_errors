from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as xlImage
from PIL import Image as pilImage

import logging, time, os


# Инициализируем логгер
logger = logging.getLogger(__name__)

timestr = time.strftime('%Y.%m.%d_%H-%M')
name_file = f'{timestr}.xlsx'

count_row = 3
result_number = 1

wb = Workbook()
wb.save(name_file)
wb = load_workbook(name_file)
ws = wb.active

ws.cell(row=1, column=1, value='№')
ws.cell(row=1, column=2, value='контекст')
ws.cell(row=1, column=3, value='предлагаемая коррекция')
ws.cell(row=1, column=4, value='адрес сайта')
ws.cell(row=1, column=5, value='scrinshot карточки')

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 30

wb.save(name_file)


def save_in_file(context, correction, link, name_scrin):

    global count_row, result_number, name_file, wb

    try:

        wb = Workbook()
        wb = load_workbook(name_file)
        ws = wb.active
        ws.cell(row=count_row, column=1, value=f'{result_number}')
        ws.cell(row=count_row, column=2, value=context)
        ws.cell(row=count_row, column=3, value=correction)
        ws.cell(row=count_row, column=4, value=link)

        img = pilImage.open(name_scrin)
        img = img.convert('RGB')  # Convert RGBA to RGB
        img.save('temp.jpg')

        logo = xlImage('temp.jpg')
        logo.height = 30
        logo.width = 100
        ws.add_image(logo, f'E{count_row}')

        count_row += 1
        result_number += 1

        time.sleep(0.5)

    except Exception as e:
        logger.exception(e)

    finally:
        wb.save(name_file)
        if os.path.isfile('temp.jpg'):
            os.remove('temp.jpg')

    logger.info(f'Результат  {name_scrin} сохранен в файл -{result_number}')

if __name__ == '__main__':
    save_in_file(context='context', correction='correction', link='link', name_scrin='image_1.jpg')